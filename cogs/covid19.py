import time
import discord
import psutil
import os
import matplotlib.pyplot as plt
import openpyxl
import asyncio
from datetime import datetime
from discord.ext import commands
from evs import default, permissions

covidlib = "./lib/covid/users/"

class covid19(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.config = default.get("config.json")
        self.process = psutil.Process(os.getpid())
        # 폴더생성
        if os.path.isdir("./lib/covid/users"):
            print("covid folder exist")
        else:
            os.makedirs("./lib/covid/users")


    # 참여
    @commands.command()
    async def 약복용알리미(self, ctx):
        embed = discord.Embed(title="약복용 알리미 시스템", description="약복용 알리미 시스템을 이용하시겠습니까?", color=0xeff0f1)
        embed.set_thumbnail(
            url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False
        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            if os.path.isfile(covidlib + str(ctx.author.id) + ".xlsx"):
                embed = discord.Embed(title="약복용알리미", description="이미 참여하셨습니다.", color=0xeff0f1)
                await ctx.send(embed=embed)
            else:
                embed = discord.Embed(title="약복용알리미",
                                      description="새로 오셨군요? " + str(ctx.author.name) + "님을 위한 파일들을 생성중이에요!",
                                      color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                await ctx.send(embed=embed)

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.cell(row=1, column=1).value = "0"  #finenum
                ws.cell(row=1, column=2).value = "0"  #time
                wb.save(covidlib + str(ctx.author.id) + ".xlsx")
                wb.close()
                embed = discord.Embed(title="약복용알리미",
                                      description=str(ctx.author.name) + " 생성 완료!",
                                      color=0xeff0f1)
                await ctx.send(embed=embed)

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="약복용알리미", description="너무 늦으셨습니다. 다음 기회에...", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="약복용 알리미 시스템", description="약복용 알리미 시스템을 이용하시겠습니까?", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await msg.edit(content=embed)

    @commands.command()
    async def 약시간(self,ctx, *, content:str):
        if os.path.isfile(covidlib + str(ctx.author.id) + ".xlsx"):
            await ctx.send("정하신 시간을 저장하는 중입니다.")
            wb = openpyxl.load_workbook(covidlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            ws.cell(row=1, column=2).value = content
            wb.save(covidlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            await ctx.send("완료")
            now = datetime.now()
            await ctx.send("현재시간:"+str(now.year) + "년" + str(now.month) + "월" + str(now.day) +"일" + str(now.hour) + "시" + str(now.minute) + "분" + str(now.second) + "초")
        else:
            embed = discord.Embed(title="약복용알리미", description="`약복용알리미`를 이용하여 먼저 참여해주세요.", color=0xeff0f1)
            await ctx.send(embed=embed)

def setup(bot):
    bot.add_cog(covid19(bot))
