import discord
import time
import psutil
import os
import asyncio
import openpyxl
import random
import math
import numpy as np
import datetime
import matplotlib.pyplot as plt

from datetime import datetime
from discord.ext import commands
from evs import default, permissions

userlib = "./lib/economy/users/"
stocklib = "./lib/economy/stocks/"
cachelib = "./lib/cache/"
categories = ["농업", "목축업", "광업", "제조업", "인프라설계업", "운송업", "언론", "금융", "방위산업", "교육", "의료", "중공업", "전자산업", "대행업", "게임", "IT", "복합"]

def keundon(value: int):
    value = int(value)
    if value < 0:
        return "변수는 음수값을 가질 수 없습니다."
    elif 0 <= value < 10000:
        return str(value)
    elif 10000 <= value < 100000000:
        return str(math.floor(value / 10000)) + "만 " + str(value - math.floor(value / 10000) * 10000)
    elif 100000000 <= value < 1000000000000:
        return str(math.floor(value / 100000000)) + "억 " + str(math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(value - math.floor(value / 10000) * 10000)
    elif 1000000000000 <= value < 10000000000000000:
        return str(math.floor(value / 1000000000000)) + "조 " + str(math.floor(value / 100000000) - math.floor(value / 1000000000000) * 10000) + "억 " + str(math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(value - math.floor(value / 10000) * 10000)
    elif 10000000000000000 <= value < 100000000000000000000:
        return str(math.floor(value / 10000000000000000)) + "경 " + str(math.floor(value / 1000000000000) - math.floor(value / 10000000000000000) * 10000) + "조 " + str(math.floor(value / 100000000) - math.floor(value / 1000000000000) * 10000) + "억 " + str(math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(value - math.floor(value / 10000) * 10000)
    else:
        return "변수의 크기가 너무 큽니다."


class economy_ko(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.config = default.get("config.json")
        self.process = psutil.Process(os.getpid())
        # 폴더생성
        if os.path.isdir("./lib/economy/users"):
            print("user folder exist")
        else:
            os.makedirs("./lib/economy/users")

        if os.path.isdir("./lib/economy/stocks"):
            print("stocks folder exist")
        else:
            os.makedirs("./lib/economy/stocks")
            
    # 메시지당 돈
    @commands.Cog.listener()
    async def on_message(self, ctx):
        if ctx.guild.id == 749595288280498188:
            if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
                randomnum = random.randrange(10, 30)
                wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
                ws = wb.active
                suvmoney = int(ws.cell(row=1, column=2).value)
                suvmoney = suvmoney + randomnum
                ws.cell(row=1, column=2).value = str(suvmoney)
                wb.save(userlib + str(ctx.author.id) + ".xlsx")
                wb.close()

    # 참여
    @commands.command()
    async def 참여(self, ctx):

        embed = discord.Embed(title="케테르 경제", description="케테르 경제에 참여하시겠습니까?", color=0xeff0f1)
        embed.set_thumbnail(
            url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False

        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
                embed = discord.Embed(title="케테르 경제", description="이미 참여하셨습니다.", color=0xeff0f1)
                await ctx.send(embed=embed)
            else:
                embed = discord.Embed(title="케테르 경제",
                                      description="새로 오셨군요? " + str(ctx.author.name) + "님을 위한 파일들을 생성중이에요!",
                                      color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                await ctx.send(embed=embed)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.cell(row=1, column=1).value = "Hello World"  #:)
                ws.cell(row=1, column=2).value = "8600000"  # money
                ws.cell(row=1, column=3).value = "0"  # pres
                ws.cell(row=1, column=4).value = "-"  # rank
                ws.cell(row=1, column=5).value = "0"  # timesleep
                ws.cell(row=1, column=6).value = "0"  # gamblesleep
                ws.cell(row=2, column=1).value = "None"  # status
                ws.cell(row=2, column=2).value = "0"  # perfect
                ws.cell(row=2, column=3).value = "0"  # great
                ws.cell(row=2, column=4).value = "0"  # good
                ws.cell(row=2, column=5).value = "0"  # bad
                ws.cell(row=3, column=1).value = "0"  # tsucc
                ws.cell(row=3, column=2).value = "0"  # tfail
                ws.cell(row=3, column=3).value = "0"  # fails
                ws.cell(row=4, column=1).value = "0"  # home count
                ws.cell(row=4, column=2).value = "[1]"  # title
                ws.cell(row=4, column=3).value = "1"  # header
                ws.cell(row=4, column=4).value = "1"  # tail
                ws.cell(row=5, column=1).value = "100"  # HP
                ws.cell(row=5, column=2).value = "100"  # STR
                ws.cell(row=5, column=3).value = "100"  # DEF
                ws.cell(row=5, column=4).value = "100"  # INT
                wb.save(userlib + str(ctx.author.id) + ".xlsx")
                wb.close()
                time.sleep(1)
                embed = discord.Embed(title="케테르 경제",
                                      description=str(ctx.author.name) + " 생성 완료!",
                                      color=0xeff0f1)
                await ctx.send(embed=embed)

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="케테르 경제", description="서명하지 않으셨습니다. 다음 기회에..", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="케테르 경제", description="케테르 경제에 참여하시겠습니까?", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await msg.edit(content=embed)

    @commands.command(aliases=['돈내놔', '돈줘'])
    async def 돈받기(self, ctx):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            num = random.randrange(100, 120)
            jackpot = random.random()
            if jackpot < 0.001:
                num = num * 100000
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            getmoney = ws.cell(row=1, column=2).value
            getmoney = int(getmoney) + int(num)
            ws.cell(row=1, column=2).value = str(getmoney)
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KET", description="<@" + str(ctx.author.id) + "> " + str(
                num) + "<:ket:753449741186105375>을 받았어요!", color=0xeff0f1)
            await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command()
    async def 돈(self, ctx):
        if (ctx.message.mentions.__len__() > 0):
            for user in ctx.message.mentions:
                if os.path.isfile(userlib + str(user.id) + ".xlsx"):
                    wb = openpyxl.load_workbook(userlib + str(user.id) + ".xlsx")
                    ws = wb.active
                    money = ws.cell(row=1, column=2).value
                    wb.close()
                    kundon = keundon(int(money))
                    embed = discord.Embed(title="KET", description="<@" + str(
                        user.id) + ">님은 " + kundon + "<:ket:753449741186105375>을 가지고 계십니다!", color=0xeff0f1)
                    await ctx.send(embed=embed)
                else:
                    embed = discord.Embed(title="NO", description="유저가 ``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
                    embed.set_thumbnail(
                        url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                    await ctx.send(embed=embed)
        else:
            if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
                wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
                ws = wb.active
                money = ws.cell(row=1, column=2).value
                wb.close()
                kundon = keundon(int(money))
                embed = discord.Embed(title="KET", description="<@" + str(
                    ctx.author.id) + "> " + kundon + "<:ket:753449741186105375>을 가지고 계십니다!", color=0xeff0f1)
                await ctx.send(embed=embed)
            else:
                embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)

    @commands.command(aliases=['프리스티지', '프레스티지', 'ㅎㅍ'])
    async def 호프(self, ctx):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            prestige = ws.cell(row=1, column=3).value
            wb.close()
            embed = discord.Embed(title="PRESTIGE", description="<@" + str(ctx.author.id) + "> " + str(
                prestige) + "<:pre:753458787465297993>을 가지고 계십니다!", color=0xeff0f1)
            await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['ㄷㅂ'])
    async def 도박(self, ctx, val: int):
        if val <= 0:
            embed = discord.Embed(title="NO", description="0 이하로는 베팅할 수 없어요.", color=0xeff0f1)
            await ctx.send(embed=embed)
            return None
        if val > 80000000000:
            embed = discord.Embed(title="NO", description="베팅금은 800억 을 초과할 수 없어요.", color=0xeff0f1)
            await ctx.send(embed=embed)
            return None
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            money = ws.cell(row=1, column=2).value
            ti = ws.cell(row=1, column=6).value
            if float(ti) > (time.time() - 60):
                next = float(ti) + 60 - round(time.time())
                embed = discord.Embed(title="NO", description="현재 **" + ctx.author.name + "**님은 도박이 불가능 합니다.", color=0xeff0f1)
                if next > 59:
                    embed.set_footer(text="다음 도박 가능까지 " + str(math.floor(next/60)) + "분 " + str(round((round(next)/60 - math.floor(next/60))*60)) + "초")
                else:
                    embed.set_footer(text="다음 도박 가능까지 " + str(round(next)) + "초")
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if int(money) > val:
                discrim = random.random()
                enjail = random.random()
                if enjail < 0.0001:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - val)
                    ws.cell(row=3, column=3).value = "0"
                    ws.cell(row=1, column=5).value = str(time.time() + 259200)
                    ws.cell(row=1, column=6).value = str(time.time() + 259200)
                    embed = discord.Embed(title="KMF", description="<@" + str(ctx.author.id) + "> 당신은 불법 도박죄로 기소되었습니다. 최종판결은 다음과 같습니다 : 징역 72시간", color=0xeff0f1)
                    return await ctx.send(embed=embed)
                if enjail < 0.0005:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - val)
                    ws.cell(row=3, column=3).value = "0"
                    ws.cell(row=1, column=5).value = str(time.time() + 86400)
                    ws.cell(row=1, column=6).value = str(time.time() + 86400)
                    embed = discord.Embed(title="KMF", description="<@" + str(ctx.author.id) + "> 당신은 불법 도박죄로 기소되었습니다. 최종판결은 다음과 같습니다 : 징역 24시간", color=0xeff0f1)
                    return await ctx.send(embed=embed)
                if enjail < 0.001:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - val)
                    ws.cell(row=3, column=3).value = "0"
                    ws.cell(row=1, column=5).value = str(time.time() + 21600)
                    ws.cell(row=1, column=6).value = str(time.time() + 21600)
                    embed = discord.Embed(title="KMF", description="<@" + str(ctx.author.id) + "> 당신은 불법 도박죄로 기소되었습니다. 최종판결은 다음과 같습니다 : 징역 6시간", color=0xeff0f1)
                    return await ctx.send(embed=embed)
                if discrim < 0.02:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + 11 * val)
                    ws.cell(row=3, column=3).value = "0"
                    embed = discord.Embed(title="도박", description="<@" + str(
                        ctx.author.id) + "> " + "축하합니다! 대박이 나서 12배를 획득 하셨어요! 🎉\n획득량:" + str(
                        12 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
                elif 0.02 < discrim < 0.05 + math.sqrt(int(ws.cell(row=3, column=3).value) * 100) / 100:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + 2 * val)
                    ws.cell(row=3, column=3).value = "0"
                    embed = discord.Embed(title="도박", description="<@" + str(
                        ctx.author.id) + "> " + "축하합니다! 도박에 성공하셔서 3배를 획득 하셨어요! 🎉\n획득량:" + str(
                        3 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
                elif 0.05 + math.sqrt(int(ws.cell(row=3, column=3).value) * 100) / 100 < discrim < 0.1 + math.sqrt(
                        int(ws.cell(row=3, column=3).value) * 100) / 50:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + val)
                    ws.cell(row=3, column=3).value = "0"
                    embed = discord.Embed(title="도박", description="<@" + str(
                        ctx.author.id) + "> " + "축하합니다! 도박에 성공하셔서 2배를 획득 하셨어요! 🎉\n획득량:" + str(
                        2 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
                else:
                    emj = "<:dar:754345236574109716>"
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - val)
                    ws.cell(row=3, column=3).value = str(int(ws.cell(row=3, column=3).value) + 1)
                    embed = discord.Embed(title="도박", description="<@" + str(
                        ctx.author.id) + "> " + "도박에 실패하여 돈을 잃으셨습니다. " + emj, color=0xeff0f1)
                ws.cell(row=1, column=6).value = str(time.time())
                wb.save(userlib + str(ctx.author.id) + ".xlsx")
                wb.close()
                await ctx.send(embed=embed)
            else:
                embed = discord.Embed(title="NO", description="보유하신 잔액보다 큰 금액을 베팅할 수는 없어요.", color=0xeff0f1)
                await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['ㅇㅇ'])
    async def 올인(self, ctx):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            val = int(ws.cell(row=1, column=2).value)
            ti = ws.cell(row=1, column=6).value
            if float(ti) > (time.time() - 60):
                next = float(ti) + 60 - round(time.time())
                embed = discord.Embed(title="NO", description="현재 **" + ctx.author.name + "**님은 도박이 불가능 합니다.", color=0xeff0f1)
                if next > 59:
                    embed.set_footer(text="다음 도박 가능까지 " + str(math.floor(next/60)) + "분 " + str(round((round(next)/60 - math.floor(next/60))*60)) + "초")
                else:
                    embed.set_footer(text="다음 도박 가능까지 " + str(round(next)) + "초")
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if val > 80000000000:
                embed = discord.Embed(title="NO", description="전재산이 800억을 초과하여 올인을 사용하실 수 없습니다.", color=0xeff0f1)
                await ctx.send(embed=embed)
                return None
            discrim = random.random()
            enjail = random.random()
            if enjail < 0.0001:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - val)
                ws.cell(row=3, column=3).value = "0"
                ws.cell(row=1, column=5).value = str(time.time() + 259200)
                ws.cell(row=1, column=6).value = str(time.time() + 259200)
                embed = discord.Embed(title="KMF", description="<@" + str(ctx.author.id) + "> 당신은 불법 도박죄로 기소되었습니다. 최종판결은 다음과 같습니다 : 징역 72시간", color=0xeff0f1)
                return await ctx.send(embed=embed)
            if enjail < 0.0005:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - val)
                ws.cell(row=3, column=3).value = "0"
                ws.cell(row=1, column=5).value = str(time.time() + 86400)
                ws.cell(row=1, column=6).value = str(time.time() + 86400)
                embed = discord.Embed(title="KMF", description="<@" + str(ctx.author.id) + "> 당신은 불법 도박죄로 기소되었습니다. 최종판결은 다음과 같습니다 : 징역 24시간", color=0xeff0f1)
                return await ctx.send(embed=embed)
            if enjail < 0.001:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - val)
                ws.cell(row=3, column=3).value = "0"
                ws.cell(row=1, column=5).value = str(time.time() + 21600)
                ws.cell(row=1, column=6).value = str(time.time() + 21600)
                embed = discord.Embed(title="KMF", description="<@" + str(ctx.author.id) + "> 당신은 불법 도박죄로 기소되었습니다. 최종판결은 다음과 같습니다 : 징역 6시간", color=0xeff0f1)
                return await ctx.send(embed=embed)
            if discrim < 0.02:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) * 12)
                ws.cell(row=3, column=3).value = "0"
                embed = discord.Embed(title="올인", description="<@" + str(
                    ctx.author.id) + "> " + "축하합니다! 대박이 나서 12배를 획득 하셨어요! 🎉\n획득량:" + str(
                    12 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
            elif 0.02 < discrim < 0.05 + math.sqrt(int(ws.cell(row=3, column=3).value) * 100) / 100:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) * 3)
                ws.cell(row=3, column=3).value = "0"
                embed = discord.Embed(title="올인", description="<@" + str(
                    ctx.author.id) + "> " + "축하합니다! 올인에 성공하셔서 3배를 획득 하셨어요! 🎉\n획득량:" + str(
                    3 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
            elif 0.05 + math.sqrt(int(ws.cell(row=3, column=3).value) * 100) / 100 < discrim < 0.1 + math.sqrt(
                    int(ws.cell(row=3, column=3).value) * 100) / 50:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) * 2)
                ws.cell(row=3, column=3).value = "0"
                embed = discord.Embed(title="올인", description="<@" + str(
                    ctx.author.id) + "> " + "축하합니다! 올인에 성공하셔서 2배를 획득 하셨어요! 🎉\n획득량:" + str(
                    2 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
            else:
                emj = "<:dar:754345236574109716>"
                ws.cell(row=1, column=2).value = "0"
                ws.cell(row=3, column=3).value = str(int(ws.cell(row=3, column=3).value) + 1)
                embed = discord.Embed(title="도박", description="올인에 실패하여 전재산을 잃으셨습니다. " + emj, color=0xeff0f1)
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command()
    async def 송금(self, ctx, mention: str, valu: int):
        if (ctx.message.mentions.__len__() > 0):
            for user in ctx.message.mentions:
                if os.path.isfile(userlib + str(user.id) + ".xlsx"):
                    wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
                    ws = wb.active
                    money = int(ws.cell(row=1, column=2).value)
                    if int(money) > valu:
                        wb2 = openpyxl.load_workbook(userlib + str(user.id) + ".xlsx")
                        ws2 = wb2.active
                        money2 = int(ws2.cell(row=1, column=2).value)
                        money2 = money2 + round(valu * 92 / 100)
                        ws2.cell(row=1, column=2).value = str(money2)
                        wb2.save(userlib + str(user.id) + ".xlsx")
                        wb2.close()
                        money = money - valu
                        ws.cell(row=1, column=2).value = str(money)
                        wb.save(userlib + str(ctx.author.id) + ".xlsx")
                        wb.close()
                        embed = discord.Embed(title="송금", description="<@" + str(ctx.author.id) + "> " + str(
                            round(valu * 92 / 100)) + " <:ket:753449741186105375>" + "송금 완료(세율 8%)", color=0xeff0f1)
                        await ctx.send(embed=embed)
                    else:
                        embed = discord.Embed(title="NO", description="보유하신 잔액보다 큰 금액을 송금할 수는 없어요.", color=0xeff0f1)
                        await ctx.send(embed=embed)
                else:
                    embed = discord.Embed(title="NO", description="유저가 ``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
                    embed.set_thumbnail(
                        url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                    await ctx.send(embed=embed)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 초기화(self, ctx):
        file_list = os.listdir(userlib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        for i in range(len(file_list)):
            wb = openpyxl.load_workbook(userlib + file_list[i])
            ws = wb.active
            if int(ws.cell(row=1, column=3).value) <= 1000:
                ws.cell(row=1, column=3).value = str(
                    int(ws.cell(row=1, column=3).value) + math.floor(int(ws.cell(row=1, column=2).value) / 1000000000))
            else:
                ws.cell(row=1, column=3).value = str(round(int(ws.cell(row=1, column=3).value) / 2) + math.ceil(
                    int(ws.cell(row=1, column=2).value) / 1000000000))
            ws.cell(row=1, column=1).value = "Hello World"  #:)
            ws.cell(row=1, column=2).value = "8600000"  # money
            ws.cell(row=1, column=4).value = "-"  # rank
            ws.cell(row=1, column=5).value = "0"  # timesleep
            ws.cell(row=1, column=6).value = "0"  # gamblesleep
            ws.cell(row=2, column=1).value = "None"  # status
            ws.cell(row=2, column=2).value = "0"  # perfect
            ws.cell(row=2, column=3).value = "0"  # great
            ws.cell(row=2, column=4).value = "0"  # good
            ws.cell(row=2, column=5).value = "0"  # bad
            for j in range(1, 101):
                ws.cell(row=6, column=j).value = None  # stocks
                ws.cell(row=7, column=j).value = None  # stocks
            wb.save(userlib + file_list[i])
            wb.close()
        embed = discord.Embed(title="Admin", description="초기화 완료", color=0xeff0f1)
        await ctx.send(embed=embed)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 돈추가(self, ctx, mention: str, value: int):
        if (ctx.message.mentions.__len__() > 0):
            for user in ctx.message.mentions:
                if os.path.isfile(userlib + str(user.id) + ".xlsx"):
                    wb = openpyxl.load_workbook(userlib + str(user.id) + ".xlsx")
                    ws = wb.active
                    money = ws.cell(row=1, column=2).value
                    money = int(money) + value
                    ws.cell(row=1, column=2).value = money
                    wb.save(userlib + str(user.id) + ".xlsx")
                    wb.close()
                    embed = discord.Embed(title="KET", description=str(money) + "<:ket:753449741186105375> 추가 완료",
                                          color=0xeff0f1)
                    await ctx.send(embed=embed)
                else:
                    embed = discord.Embed(title="NO", description="유저가 ``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
                    embed.set_thumbnail(
                        url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                    await ctx.send(embed=embed)

    @commands.command(aliases=['회사등록'])
    @commands.check(permissions.is_owner)
    async def 상장(self, ctx, name: str, stocks: int, price: int, sales: int, ratio: float, business: int):
        name = name.replace("_", " ")
        if os.path.isfile(stocklib + name + ".xlsx"):
            embed = discord.Embed(title="KMF", description="이미 상장된 기업입니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = str(int(stocks))  # 최대주
        ws.cell(row=1, column=2).value = "0"  # 매매된 주
        ws.cell(row=1, column=3).value = "1"  # 최근 거래 위치
        ws.cell(row=1, column=4).value = str(int(sales))  # 매출
        ws.cell(row=1, column=5).value = str(float(ratio))  # 수익률
        ws.cell(row=1, column=6).value = str(int(business))  # 업종
        ws.cell(row=2, column=1).value = str(int(price))  # 초기가
        for i in range(2, 101):
            ws.cell(row=2, column=i).value = str(int(price)) # 초기설정
        wb.save(stocklib + name + ".xlsx")
        wb.close()
        time.sleep(1)
        embed = discord.Embed(title="KMF", description=name + "사 상장 완료!", color=0xeff0f1)
        await ctx.send(embed=embed)

    @commands.command(aliases=['회사삭제'])
    @commands.check(permissions.is_owner)
    async def 상장폐지(self, ctx, name: str):
        name = name.replace("_", " ")
        if os.path.isfile(stocklib + name + ".xlsx"):
            os.remove(stocklib + name + ".xlsx")
            embed = discord.Embed(title="KMF", description="해당 기업을 상장폐지 하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        embed = discord.Embed(title="KMF", description=name + "는 없는 회사명입니다.", color=0xeff0f1)
        await ctx.send(embed=embed)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 업종(self, ctx):
        embed = discord.Embed(title="KMF", description="업종코드와 내용", color=0xeff0f1)
        embed.set_thumbnail(
            url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
        for i in range(0, len(categories)):
            embed.add_field(name="code : " + str(i), value=categories[i])
        await ctx.send(embed=embed)

    @commands.command(aliases=['회사정보'])
    async def 회사(self, ctx, name: str):
        name = name.replace("_", " ")
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            stoks = ws.cell(row=1, column=1).value
            sold = ws.cell(row=1, column=2).value
            last = ws.cell(row=1, column=3).value
            sales = ws.cell(row=1, column=4).value
            ratio = ws.cell(row=1, column=5).value
            business = ws.cell(row=1, column=6).value
            price = ws.cell(row=2, column=int(last)).value
            if last == "1":
                prece = ws.cell(row=2, column=100).value
            else:
                prece = ws.cell(row=2, column=int(last) - 1).value
            wb.close()
            siga = keundon(int(price) * int(stoks))
            perc = round(int(price) * 100 / int(prece) - 100, 2)
            if perc > 0:
                icon = ":small_red_triangle:"
            else:
                icon = ":small_red_triangle_down:"
            embed = discord.Embed(title=name, color=0xeff0f1)
            embed.add_field(name="시가총액", value=siga + " <:ket:753449741186105375>", inline=True)
            embed.add_field(name="주가",
                            value=keundon(price) + " <:ket:753449741186105375> (" + icon + str(abs(perc)) + "%)", inline=True)
            embed.add_field(name="매매중인 주", value=keundon(int(stoks) - int(sold)) + "주", inline=True)
            embed.add_field(name="할양된 주", value=keundon(int(sold)) + "주", inline=True)
            embed.add_field(name="매출", value=keundon(int(sales)) + " <:ket:753449741186105375>", inline=True)
            embed.add_field(name="순이익",
                            value=keundon(round(int(sales) * float(ratio) / 100)) + " <:ket:753449741186105375>", inline=True)
            embed.add_field(name="예상 배당금", value=keundon(
                round(int(sales) / int(stoks) * float(ratio) / 100)) + " <:ket:753449741186105375>", inline=True)
            embed.add_field(name="업종", value=categories[int(business)], inline=True)
            await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="해당 이름의 회사를 찾기 못하였습니다", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['회사조작'])
    @commands.check(permissions.is_owner)
    async def 주식조작(self, ctx, name: str, item: str, val: int):
        """ item 항목 : 주식총주, 주가, 매출, 수익률, 업종\n수익률의 변수 val은 10이 1%입니다. """
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
        else:
            embed = discord.Embed(title="NO", description="해당 이름의 회사를 찾지 못하였습니다", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        if item == "주식총주":
            if val <= int(ws.cell(row=1, column=2).value):
                embed = discord.Embed(title="NO", description="총수는 매매된 주보다 적은 수로 변경할 수 없습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
            else:
                ws.cell(row=1, column=1).value = str(val)
                wb.save(stocklib + name + ".xlsx")
                wb.close()
                embed = discord.Embed(title="KMF", description="해당 사(社)의 주식총수를 변경하였습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
        if item == "주가":
            last = ws.cell(row=1, column=3).value
            if last == "100":
                next = 1
            else:
                next = int(last) + 1
            ws.cell(row=1, column=3).value = str(next)
            ws.cell(row=2, column=next).value = str(val)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 사(社)의 주가를 변경하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        if item == "매출":
            ws.cell(row=1, column=4).value = str(val)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 사(社)의 매출을 변경하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        if item == "수익률":
            val = val / 10
            if val > 100:
                embed = discord.Embed(title="NO", description="수익률은 100(%)을 넘길 수 없습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
            if val <= 0:
                embed = discord.Embed(title="NO", description="수익률은 0(%)이하일 수 없습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
            ws.cell(row=1, column=5).value = str(val)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 사(社)의 수익률을 변경하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        if item == "업종":
            if val > len(categories) - 1 or val < 0:
                embed = discord.Embed(title="NO", description="변수가 잘못 설정되었습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
            ws.cell(row=1, column=6).value = str(val)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 사(社)의 업종을 변경하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        embed = discord.Embed(title="NO", description="잘못된 변수 : " + item, color=0xeff0f1)
        embed.set_thumbnail(
            url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
        await ctx.send(embed=embed)

    @commands.command(aliases=['상장사'])
    async def 회사목록(self, ctx, plist: int):
        corps = os.listdir(stocklib)
        embed = discord.Embed(title="KMF", color=0xeff0f1)
        for i in range(0 + 10 * (plist - 1), 10 + 10 * (plist - 1)):
            try:
                embed.add_field(name=str(i + 1), value=corps[i].replace(".xlsx", ""))
            except IndexError:
                return await ctx.send(embed=embed)
        await ctx.send(embed=embed)

    @commands.command(aliases=['주식그래프'])
    async def 주식(self, ctx, name: str):
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            last = ws.cell(row=1, column=3).value
            prices = []
            if last == "100":
                for i in range(1, 101):
                    prices.append(int(ws.cell(row=2, column=i).value))
            elif last == "1":
                for i in range(2, 101):
                    prices.append(int(ws.cell(row=2, column=i).value))
                prices.append(int(ws.cell(row=2, column=1).value))
            else:
                for i in range(int(last) + 1, 101):
                    prices.append(int(ws.cell(row=2, column=i).value))
                for i in range(1, int(last) + 1):
                    prices.append(int(ws.cell(row=2, column=i).value))
            plt.figure(figsize=(39, 18))
            if prices[0] < prices[99]:
                plt.step(list(range(1, 101)), prices, 'r-')
            else:
                plt.step(list(range(1, 101)), prices, 'b-')
            plt.title(name, fontsize=64)
            plt.xticks(fontsize=32)
            plt.yticks(fontsize=32)
            plt.xlabel('Recently', fontsize=44)
            plt.ylabel('Price', fontsize=44)
            plt.savefig(str(ctx.author.id) + ".png", dpi=192)
            plt.clf()
            plt.close()
            wb.close()
            await ctx.send(file=discord.File("./" + str(ctx.author.id) + ".png"))
            os.remove(str(ctx.author.id) + '.png')
        else:
            embed = discord.Embed(title="NO", description="해당 이름의 회사를 찾지 못하였습니다", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['내주식', '보유주식', '주식통장'])
    async def 보유주(self, ctx):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            inteli = ws.cell(row=5, column=4).value
            embed = discord.Embed(title="KMF", description="<@" + str(ctx.author.id) + ">님의 주식통장", color=0xeff0f1)
            for i in range(1, math.ceil(int(inteli))):
                if ws.cell(row=6, column=i).value == None:
                    pass
                else:
                    started = ws.cell(row=8, column=i).value
                    embed.add_field(name=ws.cell(row=6, column=i).value, value=ws.cell(row=7, column=i).value + "주\n최근 구매가 : " + started + " <:ket:753449741186105375>")
            wb.close()
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['매수'])
    async def 주식구매(self, ctx, name: str, amount: int):
        if amount <= 0:
            embed = discord.Embed(title="NO", description="매매 주는 0주 이하일 수 없습니다.", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            money = ws.cell(row=1, column=2).value
            inteli = ws.cell(row=5, column=4).value
            ti = ws.cell(row=1, column=5).value
            wb.close()
            block = 0
            for i in range(1, math.ceil(int(inteli))):
                if ws.cell(row=6, column=i).value == name:
                    block = i
                    start = False
                elif ws.cell(row=6, column=i).value == None:
                    block = i
                    start = True
            if float(ti) > (time.time() - 360):
                next = float(ti) + 360 - round(time.time())
                embed = discord.Embed(title="NO", description="주식 거래 후 6분 동안은 추가 매매가 불가능 합니다.", color=0xeff0f1)
                if next > 59:
                    embed.set_footer(text="다음 주식 거래 허가까지 " + str(math.floor(next/60)) + "분 " + str(round((round(next)/60 - math.floor(next/60))*60)) + "초")
                else:
                    embed.set_footer(text="다음 주식 거래 허가까지 " + str(round(next)) + "초")
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if block == 0:
                embed = discord.Embed(title="NO", description="보유할 수 있는 주식의 종류를 넘었어요..", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            stocks = ws.cell(row=1, column=1).value
            sold = ws.cell(row=1, column=2).value
            last = ws.cell(row=1, column=3).value
            price = int(ws.cell(row=2, column=int(last)).value)
            if int(stocks) - int(sold) < amount:
                embed = discord.Embed(title="NO", description="구매하려는 주가 남지 않았습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                wb.close()
                return await ctx.send(embed=embed)
            if int(money) < price*amount:
                embed = discord.Embed(title="NO", description="돈이 부족합니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                wb.close()
                return await ctx.send(embed=embed)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + amount)
            if last == "100":
                ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(0.995 + (amount**0.2)/100 + (random.random()-0.5)/800)))
                ws.cell(row=1, column=3).value = "1"
            else:
                ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(0.995 + (amount**0.2)/100 + random.random()/800)))
                ws.cell(row=1, column=3).value = str(int(last) + 1)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            ws.cell(row=6, column=block).value = name
            ws.cell(row=8, column=block).value = str(price)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - price*amount)
            ws.cell(row=1, column=5).value = str(time.time())
            if start == True:
                ws.cell(row=7, column=block).value = str(amount)
            else:
                ws.cell(row=7, column=block).value = str(int(ws.cell(row=7, column=block).value) + amount)
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 주를 " + str(amount) + "주 만큼 구매하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['매각'])
    async def 주식판매(self, ctx, name: str, amount: int):
        if amount <= 0:
            embed = discord.Embed(title="NO", description="매매 주는 0주 이하일 수 없습니다.", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            inteli = ws.cell(row=5, column=4).value
            ti = ws.cell(row=1, column=5).value
            wb.close()
            block = 0
            for i in range(1, math.ceil(int(inteli))):
                if ws.cell(row=6, column=i).value == name:
                    block = i
                    start = False
            if block == 0:
                embed = discord.Embed(title="NO", description="해당 이름의 주식을 보유하고 계시지 않아요..", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if amount > int(ws.cell(row=7, column=block).value):
                embed = discord.Embed(title="NO", description="매각하려는 주만큼을 보유하고 계시지 않습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if float(ti) > (time.time() - 360):
                next = float(ti) + 360 - round(time.time())
                embed = discord.Embed(title="NO", description="주식 거래 후 6분 동안은 추가 매매가 불가능 합니다.", color=0xeff0f1)
                if next > 59:
                    embed.set_footer(text="다음 주식 거래 허가까지 " + str(math.floor(next/60)) + "분 " + str(round((round(next)/60 - math.floor(next/60))*60)) + "초")
                else:
                    embed.set_footer(text="다음 주식 거래 허가까지 " + str(round(next)) + "초")
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            last = ws.cell(row=1, column=3).value
            price = int(ws.cell(row=2, column=int(last)).value)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - amount)
            if last == "100":
                ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(1.01 - (amount**0.2)/100 + (random.random()-0.5)/800)))
                ws.cell(row=1, column=3).value = "1"
            else:
                ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(1.01 - (amount**0.2)/100 + random.random()/800)))
                ws.cell(row=1, column=3).value = str(int(last) + 1)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            if amount == int(ws.cell(row=7, column=block).value):
                ws.cell(row=6, column=block).value = None
                ws.cell(row=7, column=block).value = None
                ws.cell(row=8, column=block).value = None
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + round(price*amount*0.94))
                ws.cell(row=1, column=5).value = str(time.time())
            else:
                ws.cell(row=6, column=block).value = name
                ws.cell(row=7, column=block).value = str(int(ws.cell(row=7, column=block).value) - amount)
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + round(price*amount*0.96))
                ws.cell(row=1, column=5).value = str(time.time())
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 주를 " + str(amount) + "주 만큼 매각하였습니다.", color=0xeff0f1)
            embed.add_field(name="판매가", value=keundon(amount*price) + " <:ket:753449741186105375> (세율 4% : " + keundon(round(amount*price*0.04)) + " <:ket:753449741186105375>)")
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['최매수'])
    async def 최대주식구매(self, ctx, name: str):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            assets = int(ws.cell(row=1, column=2).value)
            wb.close()
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            money = ws.cell(row=1, column=2).value
            inteli = ws.cell(row=5, column=4).value
            ti = ws.cell(row=1, column=5).value
            wb.close()
            block = 0
            for i in range(1, math.ceil(int(inteli))):
                if ws.cell(row=6, column=i).value == name:
                    block = i
                    start = False
                elif ws.cell(row=6, column=i).value == None:
                    block = i
                    start = True
            if float(ti) > (time.time() - 360):
                next = float(ti) + 360 - round(time.time())
                embed = discord.Embed(title="NO", description="주식 거래 후 6분 동안은 추가 매매가 불가능 합니다.", color=0xeff0f1)
                if next > 59:
                    embed.set_footer(text="다음 주식 거래 허가까지 " + str(math.floor(next/60)) + "분 " + str(round((round(next)/60 - math.floor(next/60))*60)) + "초")
                else:
                    embed.set_footer(text="다음 주식 거래 허가까지 " + str(round(next)) + "초")
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if block == 0:
                embed = discord.Embed(title="NO", description="보유할 수 있는 주식의 종류를 넘었어요..", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            stocks = ws.cell(row=1, column=1).value
            sold = ws.cell(row=1, column=2).value
            last = ws.cell(row=1, column=3).value
            price = int(ws.cell(row=2, column=int(last)).value)
            amount = math.floor(assets/price)
            if amount <= 0:
                embed = discord.Embed(title="NO", description="매매 주는 0주 이하일 수 없습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if int(stocks) - int(sold) < amount:
                embed = discord.Embed(title="NO", description="구매하려는 주가 남지 않았습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                wb.close()
                return await ctx.send(embed=embed)
            if int(money) < price*amount:
                embed = discord.Embed(title="NO", description="돈이 부족합니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                wb.close()
                return await ctx.send(embed=embed)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + amount)
            if last == "100":
                ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(0.995 + (amount**0.2)/100 + (random.random()-0.5)/800)))
                ws.cell(row=1, column=3).value = "1"
            else:
                ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(0.995 + (amount**0.2)/100 + random.random()/800)))
                ws.cell(row=1, column=3).value = str(int(last) + 1)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            ws.cell(row=6, column=block).value = name
            ws.cell(row=8, column=block).value = str(price)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - price*amount)
            ws.cell(row=1, column=5).value = str(time.time())
            if start == True:
                ws.cell(row=7, column=block).value = str(amount)
            else:
                ws.cell(row=7, column=block).value = str(int(ws.cell(row=7, column=block).value) + amount)
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 주를 " + str(amount) + "주 만큼 구매하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['전매각'])
    async def 일괄주식판매(self, ctx, name: str):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            inteli = ws.cell(row=5, column=4).value
            ti = ws.cell(row=1, column=5).value
            wb.close()
            block = 0
            for i in range(1, math.ceil(int(inteli))):
                if ws.cell(row=6, column=i).value == name:
                    block = i
                    start = False
            if block == 0:
                embed = discord.Embed(title="NO", description="해당 이름의 주식을 보유하고 계시지 않아요..", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if float(ti) > (time.time() - 360):
                next = float(ti) + 360 - round(time.time())
                embed = discord.Embed(title="NO", description="주식 거래 후 6분 동안은 추가 매매가 불가능 합니다.", color=0xeff0f1)
                if next > 59:
                    embed.set_footer(text="다음 주식 거래 허가까지 " + str(math.floor(next/60)) + "분 " + str(round((round(next)/60 - math.floor(next/60))*60)) + "초")
                else:
                    embed.set_footer(text="다음 주식 거래 허가까지 " + str(round(next)) + "초")
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            amount = int(ws.cell(row=7, column=block).value)
            wb.close()
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            last = ws.cell(row=1, column=3).value
            price = int(ws.cell(row=2, column=int(last)).value)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - amount)
            if last == "100":
                ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(1.01 - (amount**0.2)/100 + (random.random()-0.5)/800)))
                ws.cell(row=1, column=3).value = "1"
            else:
                ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(1.01 - (amount**0.2)/100 + random.random()/800)))
                ws.cell(row=1, column=3).value = str(int(last) + 1)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            ws.cell(row=6, column=block).value = None
            ws.cell(row=7, column=block).value = None
            ws.cell(row=8, column=block).value = None
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + round(price*amount*0.96))
            ws.cell(row=1, column=5).value = str(time.time())
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 주를 " + str(amount) + "주 만큼 매각하였습니다.", color=0xeff0f1)
            embed.add_field(name="판매가", value=keundon(amount*price) + " <:ket:753449741186105375> (세율 4% : " + keundon(round(amount*price*0.04)) + " <:ket:753449741186105375>)")
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['관매수'])
    @commands.check(permissions.is_owner)
    async def 어드민주식(self, ctx, name: str, amount: int):
        if amount <= 0:
            embed = discord.Embed(title="NO", description="매매 주는 0주 이하일 수 없습니다.", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            money = ws.cell(row=1, column=2).value
            inteli = ws.cell(row=5, column=4).value
            wb.close()
            block = 0
            for i in range(1, math.ceil(int(inteli))):
                if ws.cell(row=6, column=i).value == name:
                    block = i
                    start = False
                elif ws.cell(row=6, column=i).value == None:
                    block = i
                    start = True
            if block == 0:
                embed = discord.Embed(title="NO", description="보유할 수 있는 주식의 종류를 넘었어요..", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            stocks = ws.cell(row=1, column=1).value
            sold = ws.cell(row=1, column=2).value
            last = ws.cell(row=1, column=3).value
            price = int(ws.cell(row=2, column=int(last)).value)
            if int(stocks) - int(sold) < amount:
                embed = discord.Embed(title="NO", description="구매하려는 주가 남지 않았습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                wb.close()
                return await ctx.send(embed=embed)
            if int(money) < price*amount:
                embed = discord.Embed(title="NO", description="돈이 부족합니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                wb.close()
                return await ctx.send(embed=embed)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + amount)
            if last == "100":
                ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(0.995 + (amount**0.2)/100 + (random.random()-0.5)/800)))
                ws.cell(row=1, column=3).value = "1"
            else:
                ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(0.995 + (amount**0.2)/100 + random.random()/800)))
                ws.cell(row=1, column=3).value = str(int(last) + 1)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            ws.cell(row=6, column=block).value = name
            ws.cell(row=8, column=block).value = str(price)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - price*amount)
            ws.cell(row=1, column=5).value = str(time.time())
            if start == True:
                ws.cell(row=7, column=block).value = str(amount)
            else:
                ws.cell(row=7, column=block).value = str(int(ws.cell(row=7, column=block).value) + amount)
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 주를 " + str(amount) + "주 만큼 구매하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['관매각'])
    @commands.check(permissions.is_owner)
    async def 어드민주식판매(self, ctx, name: str, amount: int):
        if amount <= 0:
            embed = discord.Embed(title="NO", description="매매 주는 0주 이하일 수 없습니다.", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            inteli = ws.cell(row=5, column=4).value
            wb.close()
            block = 0
            for i in range(1, math.ceil(int(inteli))):
                if ws.cell(row=6, column=i).value == name:
                    block = i
                    start = False
            if amount > int(ws.cell(row=7, column=block).value):
                embed = discord.Embed(title="NO", description="매각하려는 주만큼을 보유하고 계시지 않습니다.", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
            if block == 0:
                embed = discord.Embed(title="NO", description="해당 이름의 주식을 보유하고 계시지 않아요..", color=0xeff0f1)
                embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                return await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
            embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            return await ctx.send(embed=embed)
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            last = ws.cell(row=1, column=3).value
            price = int(ws.cell(row=2, column=int(last)).value)
            ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - amount)
            if last == "100":
                ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(1.01 - (amount**0.2)/100 + (random.random()-0.5)/800)))
                ws.cell(row=1, column=3).value = "1"
            else:
                ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(1.01 - (amount**0.2)/100 + random.random()/800)))
                ws.cell(row=1, column=3).value = str(int(last) + 1)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            if amount == int(ws.cell(row=7, column=block).value):
                ws.cell(row=6, column=block).value = None
                ws.cell(row=7, column=block).value = None
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + round(price*amount*0.94))
                ws.cell(row=1, column=5).value = str(time.time())
            else:
                ws.cell(row=6, column=block).value = name
                ws.cell(row=7, column=block).value = str(int(ws.cell(row=7, column=block).value) - amount)
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + round(price*amount*0.96))
                ws.cell(row=1, column=5).value = str(time.time())
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 주를 " + str(amount) + "주 만큼 매각하였습니다.", color=0xeff0f1)
            embed.add_field(name="판매가", value=keundon(amount*price) + " <:ket:753449741186105375> (세율 4% : " + keundon(round(amount*price*0.04)) + " <:ket:753449741186105375>)")
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            
    @commands.command()
    async def 테스트그래프(self, ctx):
        x = np.linspace(-6, 6, 30)
        y = np.linspace(-6, 6, 30)
        x, y = np.meshgrid(x, y)
        z = np.sin(np.sqrt(x**2 + y**2))

        fig = plt.figure(figsize=(12, 6))
        ax = plt.axes(projection='3d')
        ax.contour3D(x, y, z, 20, cmap=plt.cm.rainbow)
        plt.savefig(str(ctx.author.id) + ".png", dpi=192)
        plt.title("ax.contour3D")
        plt.clf()
        plt.close()
        await ctx.send(file=discord.File("./" + str(ctx.author.id) + ".png"))
        os.remove(str(ctx.author.id) + '.png')

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 전체초기화(self, ctx):
        file_list = os.listdir(userlib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        for i in range(len(file_list)):
            os.remove(userlib + file_list[i])
            await ctx.send(file_list[i] + "deleted")

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 상장초기화(self, ctx):
        file_list = os.listdir(stocklib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        for i in range(len(file_list)):
            os.remove(stocklib + file_list[i])
            await ctx.send(file_list[i] + "deleted")

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 주가초기화(self, ctx):
        f = open(cachelib + "is_started.ccf", "w")
        file_list = os.listdir(stocklib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        cycles = 0
        if not os.path.isfile(cachelib + "is_started.ccf"):
            return await ctx.send(str(cycles) + "cycle stopped")
        for i in range(len(file_list)):
            wb = openpyxl.load_workbook(stocklib + file_list[i])
            ws = wb.active
            last = ws.cell(row=1, column=3).value
            if last == "100":
                ws.cell(row=2, column=1).value = str(round(float(ws.cell(row=1, column=4).value)*float(ws.cell(row=1, column=5).value)*10/float(ws.cell(row=1, column=2).value)))
                ws.cell(row=1, column=3).value = "1"
            else:
                ws.cell(row=2, column=int(last) + 1).value = str(round(float(ws.cell(row=1, column=4).value)*float(ws.cell(row=1, column=5).value)*10/float(ws.cell(row=1, column=2).value)))
                ws.cell(row=1, column=3).value = str(int(last) + 1)
            wb.save(stocklib + file_list[i])
            wb.close()

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 주가변동(self, ctx, cycle :int):
        if os.path.isfile(cachelib + "is_started.ccf"):
            return await ctx.send("이미 실행중입니다.")
        await ctx.send("코드를 실행합니다.")
        f = open(cachelib + "is_started.ccf", "w")
        f.close()
        file_list = os.listdir(stocklib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        cycles = 0
        while cycles < cycle:
            cycles += 1
            if not os.path.isfile(cachelib + "is_started.ccf"):
                return await ctx.send(str(cycles) + "cycle stopped")
            for i in range(len(file_list)):
                wb = openpyxl.load_workbook(stocklib + file_list[i])
                ws = wb.active
                last = ws.cell(row=1, column=3).value
                if last == "100":
                    ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(0.995 + (random.random()-0.5)/20)))
                    ws.cell(row=1, column=3).value = "1"
                else:
                    ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(0.995 + (random.random()-0.5)/20)))
                    ws.cell(row=1, column=3).value = str(int(last) + 1)
                wb.save(stocklib + file_list[i])
                wb.close()
            if cycles == cycle:
                await ctx.send("last cycle reseted")
                os.remove(cachelib + "is_started.ccf")
            else:
                await ctx.send(str(cycles) + "cycle reseted")
            await asyncio.sleep(300)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 불황변동(self, ctx, cycle :int):
        if os.path.isfile(cachelib + "is_started.ccf"):
            return await ctx.send("이미 실행중입니다.")
        await ctx.send("코드를 실행합니다.")
        f = open(cachelib + "is_started.ccf", "w")
        f.close()
        file_list = os.listdir(stocklib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        cycles = 0
        while cycles < cycle:
            cycles += 1
            if not os.path.isfile(cachelib + "is_started.ccf"):
                return await ctx.send(str(cycles) + "cycle stopped")
            for i in range(len(file_list)):
                wb = openpyxl.load_workbook(stocklib + file_list[i])
                ws = wb.active
                last = ws.cell(row=1, column=3).value
                if last == "100":
                    ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(0.96 + (random.random()-0.5)/20)))
                    ws.cell(row=1, column=3).value = "1"
                else:
                    ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(0.96 + (random.random()-0.5)/20)))
                    ws.cell(row=1, column=3).value = str(int(last) + 1)
                wb.save(stocklib + file_list[i])
                wb.close()
            if cycles == cycle:
                await ctx.send("last cycle reseted")
                os.remove(cachelib + "is_started.ccf")
            else:
                await ctx.send(str(cycles) + "cycle reseted")
            await asyncio.sleep(300)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 호황변동(self, ctx, cycle :int):
        if os.path.isfile(cachelib + "is_started.ccf"):
            return await ctx.send("이미 실행중입니다.")
        await ctx.send("코드를 실행합니다.")
        f = open(cachelib + "is_started.ccf", "w")
        f.close()
        file_list = os.listdir(stocklib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        cycles = 0
        while cycles < cycle:
            cycles += 1
            if not os.path.isfile(cachelib + "is_started.ccf"):
                return await ctx.send(str(cycles) + "cycle stopped")
            for i in range(len(file_list)):
                wb = openpyxl.load_workbook(stocklib + file_list[i])
                ws = wb.active
                last = ws.cell(row=1, column=3).value
                if last == "100":
                    ws.cell(row=2, column=1).value = str(round(int(ws.cell(row=2, column=100).value)*(1.005 + (random.random()-0.5)/20)))
                    ws.cell(row=1, column=3).value = "1"
                else:
                    ws.cell(row=2, column=int(last) + 1).value = str(round(int(ws.cell(row=2, column=int(last)).value)*(1.005 + (random.random()-0.5)/20)))
                    ws.cell(row=1, column=3).value = str(int(last) + 1)
                wb.save(stocklib + file_list[i])
                wb.close()
            if cycles == cycle:
                await ctx.send("last cycle reseted")
                os.remove(cachelib + "is_started.ccf")
            else:
                await ctx.send(str(cycles) + "cycle reseted")
            await asyncio.sleep(300)
    
    @commands.command()
    @commands.check(permissions.is_owner)
    async def AI변동(self, ctx, cycle :int):
        if os.path.isfile(cachelib + "is_started.ccf"):
            return await ctx.send("이미 실행중입니다.")
        await ctx.send("추가 예정입니다.")

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 변동픽스(self, ctx):
        if os.path.isfile(cachelib + "is_started.ccf"):
            os.remove(cachelib + "is_started.ccf")
            await ctx.send("캐시를 제거하였습니다.")
        else:
            await ctx.send("캐시파일이 없습니다")

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 배당시작(self, ctx):
        if os.path.isfile(stocklib + "is_divided.ccf"):
            return await ctx.send("이미 실행중입니다.")
        await ctx.send("코드를 실행합니다.")
        f = open(cachelib + "is_divided.ccf", "w")
        f.close()
        cycles = True
        while cycles == True:
            cycles = os.path.isfile(cachelib + "is_divided.ccf")
            file_list = os.listdir(stocklib)
            file_list = [file for file in file_list if file.endswith(".xlsx")]
            for i in range(len(file_list)):
                wb = openpyxl.load_workbook(stocklib + file_list[i])
                ws = wb.active
                ws.cell(row=1, column=4).value = str(round(float(ws.cell(row=1, column=4).value) * (1 + (random.random() - 0.48)/32)))
                wb.save(stocklib + file_list[i])
                wb.close()
            await ctx.send("all the stocks have been reseted")
            file_list = os.listdir(userlib)
            file_list = [file for file in file_list if file.endswith(".xlsx")]
            for i in range(len(file_list)):
                wb = openpyxl.load_workbook(userlib + file_list[i])
                ws = wb.active
                money = float(ws.cell(row=1, column=2).value)
                inteli = ws.cell(row=5, column=4).value
                for j in range(1, math.ceil(int(inteli))):
                    name = ws.cell(row=6, column=j).value
                    try:
                        amount = int(ws.cell(row=7, column=j).value)
                    except:
                        continue
                    sb = openpyxl.load_workbook(stocklib + name + ".xlsx")
                    ss = sb.active
                    percap = int(ss.cell(row=1, column=4).value) / int(ss.cell(row=1, column=1).value)
                    money += amount*percap
                    sb.close()
                ws.cell(row=1, column=2).value = str(round(money))
                wb.save(userlib + file_list[i])
                wb.close()
            await ctx.send("all the user got the dividend")
            await asyncio.sleep(86400)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 배당픽스(self, ctx):
        if os.path.isfile(cachelib + "is_divided.ccf"):
            os.remove(cachelib + "is_divided.ccf")
            await ctx.send("캐시를 제거하였습니다.")
        else:
            await ctx.send("캐시파일이 없습니다")

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 할양초기화(self, ctx):
        file_list = os.listdir(stocklib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        for i in range(len(file_list)):
            wb = openpyxl.load_workbook(stocklib + file_list[i])
            ws = wb.active
            stocks = ws.cell(row=1, column=2).value = "0"
            wb.save(stocklib + file_list[i])
            wb.close()
            await ctx.send(file_list[i] + " reseted")

def setup(bot):
    bot.add_cog(economy_ko(bot))
